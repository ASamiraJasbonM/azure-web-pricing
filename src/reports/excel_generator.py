"""
Excel report generation
"""

from typing import Dict, List, Any, Optional
from pathlib import Path
from datetime import datetime

# Try to import openpyxl
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelReportGenerator:
    """Generate Excel reports for cost estimation"""
    
    def __init__(self, output_dir: str = "output/reports"):
        """
        Initialize Excel generator
        
        Args:
            output_dir: Directory to save reports
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        if HAS_OPENPYXL:
            self._header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            self._header_font = Font(bold=True, color="FFFFFF", size=12)
            self._border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def generate_report(
        self,
        cost_data: Dict[str, Any],
        title: str = "Azure Cost Estimation"
    ) -> str:
        """
        Generate an Excel report
        
        Args:
            cost_data: Cost breakdown data
            title: Report title
            
        Returns:
            Path to saved Excel file
        """
        if not HAS_OPENPYXL:
            return "Error: openpyxl not installed. Install with: pip install openpyxl"
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(ws_summary, cost_data)
        
        # Services sheet
        ws_services = wb.create_sheet("Services")
        self._create_services_sheet(ws_services, cost_data)
        
        # Details sheet
        ws_details = wb.create_sheet("Details")
        self._create_details_sheet(ws_details, cost_data)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"azure_cost_report_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        # Save
        wb.save(filepath)
        
        return str(filepath)
    
    def _create_summary_sheet(self, ws, cost_data: Dict[str, Any]):
        """Create summary sheet"""
        # Title
        ws['A1'] = "Azure Cost Estimation Report"
        ws['A1'].font = Font(bold=True, size=16, color="0078D4")
        
        # Date
        ws['A3'] = "Generated:"
        ws['A3'].font = Font(bold=True)
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Summary data
        ws['A5'] = "COST SUMMARY"
        ws['A5'].font = Font(bold=True, size=14)
        
        summary = cost_data.get("summary", {})
        
        row = 7
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        
        for col in ['A', 'B']:
            cell = ws[f'{col}{row}']
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        metrics = [
            ("Total Monthly Cost", f"${summary.get('total_monthly_cost_usd', 0):.2f}"),
            ("Total Hourly Cost", f"${summary.get('total_hourly_cost_usd', 0):.2f}"),
            ("Currency", summary.get("currency", "USD")),
            ("Region", summary.get("region", "N/A")),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_services_sheet(self, ws, cost_data: Dict[str, Any]):
        """Create services breakdown sheet"""
        # Title
        ws['A1'] = "Service Breakdown"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Headers
        row = 3
        ws['A3'] = "Service"
        ws['B3'] = "Monthly Cost"
        ws['C3'] = "Hourly Cost"
        
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}3']
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row = 4
        for service, details in cost_data.items():
            if service == "summary":
                continue
            
            if isinstance(details, dict):
                monthly = details.get("total_monthly", 0)
                hourly = monthly / 730 if monthly else 0
                
                ws[f'A{row}'] = service
                ws[f'B{row}'] = monthly
                ws[f'C{row}'] = hourly
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_details_sheet(self, ws, cost_data: Dict[str, Any]):
        """Create detailed breakdown sheet"""
        # Title
        ws['A1'] = "Detailed Breakdown"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Headers
        headers = ["Service", "Property", "Value"]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row = 4
        for service, details in cost_data.items():
            if service == "summary":
                continue
            
            if isinstance(details, dict):
                for key, value in details.items():
                    if key != "total_monthly":
                        ws[f'A{row}'] = service
                        ws[f'B{row}'] = key
                        ws[f'C{row}'] = str(value)
                        row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
    
    def add_chart_sheet(
        self,
        excel_path: str,
        chart_data: Dict[str, float],
        chart_name: str = "Cost Chart"
    ) -> str:
        """
        Add a chart to Excel file
        
        Args:
            excel_path: Path to Excel file
            chart_data: Data for chart
            chart_name: Name of chart
            
        Returns:
            Updated Excel path
        """
        if not HAS_OPENPYXL:
            return "Error: openpyxl not installed"
        
        # Load workbook
        wb = openpyxl.load_workbook(excel_path)
        
        # Create chart sheet
        ws = wb.create_sheet(chart_name)
        
        # Add data
        ws['A1'] = "Service"
        ws['B1'] = "Cost"
        
        row = 2
        for service, cost in chart_data.items():
            ws[f'A{row}'] = service
            ws[f'B{row}'] = cost
            row += 1
        
        # Save
        wb.save(excel_path)
        
        return excel_path


# Singleton
_excel_generator: Optional[ExcelReportGenerator] = None


def get_excel_generator() -> ExcelReportGenerator:
    """Get Excel generator instance"""
    global _excel_generator
    if _excel_generator is None:
        _excel_generator = ExcelReportGenerator()
    return _excel_generator